import os
import tkinter as tk
from tkinter import messagebox, filedialog
from datetime import datetime
from openpyxl import load_workbook
from num2words import num2words
import subprocess
import tempfile

def find_soffice_path():
    possible_paths = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"
    ]
    for path in possible_paths:
        if os.path.exists(path):
            return path
    return None

def calculate_service_period(*args):
    try:
        doj = datetime.strptime(doj_var.get(), "%d-%m-%Y")
        dol = datetime.strptime(dol_var.get(), "%d-%m-%Y")

        delta_years = dol.year - doj.year
        delta_months = dol.month - doj.month
        delta_days = dol.day - doj.day

        if delta_days < 0:
            delta_months -= 1
        if delta_months < 0:
            delta_years -= 1
            delta_months += 12

        service_var.set(f"{delta_years} years - {delta_months} months")

        if delta_years >= 4:
            try:
                basic = float(basic_var.get())
                years_for_gratuity = delta_years
                if delta_months >= 6:
                    years_for_gratuity += 1

                if years_for_gratuity >= 5:
                    gratuity = int(round((basic / 26) * 15 * years_for_gratuity))
                    gratuity_var.set(str(gratuity))
                else:
                    gratuity_var.set("")
            except:
                gratuity_var.set("")
        else:
            gratuity_var.set("")
    except:
        service_var.set("")
        gratuity_var.set("")

    calculate_total()

def calculate_total(*args):
    try:
        total = (
            float(unpaid_var.get() or 0) +
            float(exgratia_var.get() or 0) +
            float(gratuity_var.get() or 0) +
            float(bonus_var.get() or 0) +
            float(leave_var.get() or 0) +
            float(others_var.get() or 0)
            - float(advance_var.get() or 0)
            - float(notice_var.get() or 0)
        )
        total_var.set(str(round(total)))
    except:
        total_var.set("")

def clear_all_fields():
    for var in [
        name_var, ref_id_var, doj_var, dol_var, service_var,
        gross_var, basic_var, unpaid_var, exgratia_var, gratuity_var,
        bonus_var, leave_var, others_var, advance_var, notice_var,
        table_leader_var, hr_var, total_var
    ]:
        var.set("")

def export_to_pdf():
    name = name_var.get().strip()
    ref_id = ref_id_var.get().strip()

    doj = doj_var.get()
    dol = dol_var.get()
    service = service_var.get()
    gross = gross_var.get()
    basic = basic_var.get()

    unpaid = unpaid_var.get()
    exgratia = exgratia_var.get()
    gratuity = gratuity_var.get()
    bonus = bonus_var.get()
    leave = leave_var.get()
    others = others_var.get()
    advance = advance_var.get()
    notice = notice_var.get()
    table_leader = table_leader_var.get()
    hr = hr_var.get()

    required = [name, doj, dol, gross, basic]
    if not all(required):
        messagebox.showerror("Error", "Please fill all mandatory fields.")
        return

    try:
        total = float(total_var.get())
        total_int = int(round(total))
        amount_words = num2words(total_int, lang='en_IN').replace("euro", "rupees").replace("cents", "paise").title()
        amount_words = f"Rupees {amount_words} Only"
    except Exception as e:
        messagebox.showerror("Error", f"Invalid number input:\n{e}")
        return

    template_path = os.path.abspath("FNF_Format.xlsx")
    if not os.path.exists(template_path):
        messagebox.showerror("Error", f"Template not found:\n{template_path}")
        return

    save_pdf_path = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf")],
        initialfile=f"{ref_id} - {name}.pdf",
        title="Save PDF As"
    )
    if not save_pdf_path:
        return

    temp_excel = os.path.join(tempfile.gettempdir(), "temp_fnf.xlsx")

    try:
        wb = load_workbook(template_path)
        ws = wb["Sheet1"]

        ws["E4"].value = name
        ws["E6"].value = ref_id
        ws["E8"].value = doj
        ws["E10"].value = dol
        ws["E12"].value = service
        ws["E14"].value = gross
        ws["E16"].value = basic
        ws["F21"].value = unpaid
        ws["F23"].value = exgratia
        ws["F25"].value = gratuity
        ws["F28"].value = bonus
        ws["F31"].value = leave
        ws["F34"].value = others
        ws["F36"].value = advance
        ws["F37"].value = notice
        ws["F40"].value = total_int
        ws["C42"].value = amount_words
        ws["B47"].value = f"Received a sum of Rs{total_int} /- {amount_words}"
        ws["C55"].value = f"TL: {table_leader}"
        ws["C57"].value = f"HR: {hr}"

        wb.save(temp_excel)

        soffice = find_soffice_path()
        if not soffice:
            messagebox.showerror("Error", "LibreOffice not found.\nPlease install LibreOffice to export PDF.")
            return

        subprocess.run([
            soffice,
            "--headless", "--convert-to", "pdf", temp_excel,
            "--outdir", os.path.dirname(save_pdf_path)
        ], check=True)

        generated_pdf = os.path.join(
            os.path.dirname(save_pdf_path),
            os.path.splitext(os.path.basename(temp_excel))[0] + ".pdf"
        )
        os.replace(generated_pdf, save_pdf_path)
        os.remove(temp_excel)

        messagebox.showinfo("Success", f"PDF saved to:\n{save_pdf_path}")
        os.startfile(os.path.dirname(save_pdf_path))

    except Exception as e:
        messagebox.showerror("Export Error", f"PDF generation failed:\n{e}")
        if os.path.exists(temp_excel):
            os.remove(temp_excel)

# ---------- UI ----------
root = tk.Tk()
root.title("Full & Final Settlement Generator")
root.geometry("620x880")

def create_labeled_entry(label, row, var, entry_type='entry'):
    tk.Label(root, text=label).grid(row=row, column=0, sticky="w", padx=10, pady=5)
    state = "readonly" if entry_type == 'readonly' else "normal"
    entry = tk.Entry(root, textvariable=var, width=30, state=state)
    entry.grid(row=row, column=1, padx=10, pady=5)
    return entry

# Variables
name_var = tk.StringVar()
ref_id_var = tk.StringVar()
doj_var = tk.StringVar()
dol_var = tk.StringVar()
service_var = tk.StringVar()
gross_var = tk.StringVar()
basic_var = tk.StringVar()
unpaid_var = tk.StringVar()
exgratia_var = tk.StringVar()
gratuity_var = tk.StringVar()
bonus_var = tk.StringVar()
leave_var = tk.StringVar()
others_var = tk.StringVar()
advance_var = tk.StringVar()
notice_var = tk.StringVar()
table_leader_var = tk.StringVar()
hr_var = tk.StringVar()
total_var = tk.StringVar()

# Inputs
create_labeled_entry("Employee Name *", 0, name_var)
create_labeled_entry("Reference ID", 1, ref_id_var)
create_labeled_entry("Date of Joining *(dd-mm-yyyy)", 2, doj_var)
create_labeled_entry("Date of Leaving *(dd-mm-yyyy)", 3, dol_var)
create_labeled_entry("Service Period", 4, service_var, 'readonly')
create_labeled_entry("Gross Salary *", 5, gross_var)
create_labeled_entry("Basic Salary *", 6, basic_var)

tk.Label(root, text="--- Payment Details ---", font=("Arial", 10, "bold")).grid(row=7, column=0, columnspan=2, pady=10)
create_labeled_entry("Unpaid Salary", 8, unpaid_var)
create_labeled_entry("Ex-Gratia", 9, exgratia_var)
create_labeled_entry("Gratuity", 10, gratuity_var, 'readonly')
create_labeled_entry("Bonus", 11, bonus_var)
create_labeled_entry("Leave Encashment", 12, leave_var)
create_labeled_entry("Other Payments", 13, others_var)
create_labeled_entry("Advance Deduction", 14, advance_var)
create_labeled_entry("Notice Pay Recovery", 15, notice_var)

create_labeled_entry("Table Leader", 16, table_leader_var)
create_labeled_entry("HR", 17, hr_var)
create_labeled_entry("Total Amount", 18, total_var, 'readonly')

# Buttons
tk.Button(
    root, text="Generate Full & Final PDF", command=export_to_pdf,
    bg="green", fg="white", font=("Arial", 12)
).grid(row=19, column=0, pady=20, padx=10)

tk.Button(
    root, text="Clear All", command=clear_all_fields,
    bg="gray", fg="white", font=("Arial", 12)
).grid(row=19, column=1, pady=20, padx=10)

# Traces for auto-calculation
doj_var.trace_add("write", calculate_service_period)
dol_var.trace_add("write", calculate_service_period)
basic_var.trace_add("write", calculate_service_period)

for var in [
    unpaid_var, exgratia_var, gratuity_var, bonus_var,
    leave_var, others_var, advance_var, notice_var
]:
    var.trace_add("write", calculate_total)

root.mainloop()
